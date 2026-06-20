from __future__ import annotations

import io
from datetime import datetime


def build_recommendation_excel(rows: list[dict], include_user_col: bool = True) -> bytes:
    """Build xlsx from recommendation rows. Returns raw bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekomendasi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    center = Alignment(horizontal="center")

    headers = []
    if include_user_col:
        headers.append("User")
    headers += ["Role Target", "Cluster", "Precision@K", "Budget Maks (Rp)", "Tanggal"]

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, row in enumerate(rows, start=2):
        col = 1
        if include_user_col:
            ws.cell(row=row_idx, column=col, value=row.get("user", "")).alignment = center
            col += 1
        ws.cell(row=row_idx, column=col, value=row.get("role_target", ""))
        col += 1
        ws.cell(row=row_idx, column=col, value=row.get("cluster", ""))
        col += 1
        ws.cell(row=row_idx, column=col, value=round(row.get("precision_at_k", 0), 4))
        col += 1
        ws.cell(row=row_idx, column=col, value=row.get("budget_max", 0))
        col += 1
        tanggal = row.get("tanggal")
        if isinstance(tanggal, datetime):
            tanggal = tanggal.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row_idx, column=col, value=tanggal or "")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_recommendation_pdf(rows: list[dict], include_user_col: bool = True) -> bytes:
    """Build PDF from recommendation rows. Returns raw bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("Laporan Rekomendasi Laptop", styles["Title"]))
    elements.append(Spacer(1, 0.5 * cm))

    headers = []
    if include_user_col:
        headers.append("User")
    headers += ["Role Target", "Cluster", "Precision@K", "Budget Maks (Rp)", "Tanggal"]

    data = [headers]
    for row in rows:
        r = []
        if include_user_col:
            r.append(str(row.get("user", "")))
        r.append(str(row.get("role_target", "")))
        r.append(str(row.get("cluster", "")))
        r.append(str(round(row.get("precision_at_k", 0), 4)))
        r.append("Rp{:,}".format(row.get("budget_max", 0)))
        tanggal = row.get("tanggal")
        if isinstance(tanggal, datetime):
            tanggal = tanggal.strftime("%Y-%m-%d")
        r.append(str(tanggal or ""))
        data.append(r)

    col_count = len(headers)
    col_width = (landscape(A4)[0] - 3 * cm) / col_count
    table = Table(data, colWidths=[col_width] * col_count)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def recommendations_to_rows(qs, include_user: bool = True) -> list[dict]:
    """Convert Recommendation queryset to list of plain dicts for export fns."""
    rows = []
    for rec in qs.select_related("user", "preference", "selected_cluster"):
        row = {
            "role_target": rec.preference.role_target if rec.preference else "",
            "cluster": rec.selected_cluster.interpretation if rec.selected_cluster else "",
            "precision_at_k": rec.precision_at_k,
            "budget_max": rec.preference.budget_max_idr if rec.preference else 0,
            "tanggal": rec.created_at,
        }
        if include_user:
            row["user"] = rec.user.username
        rows.append(row)
    return rows
